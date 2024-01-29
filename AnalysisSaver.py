"""This function writes the analysis done.

It saves all the matrices in .bin files, saves a .avi file and a file journal
wit all the information needed about the activation.
"""

import shutil
import os
import datetime
import numpy as np
import tifffile
import pyqtgraph as pg
import xlsxwriter
from openpyxl import load_workbook
from skimage.measure import regionprops

import SpotsConnection
import NucleiSpotsConnection
import WriteSptsIntsDividedByBkg


class FilteredSpotsSaver:
    """Save a parallel analysis folder with the filtered spots."""
    def __init__(self, analysis_folder, spots_3D, fnames, green4D, solidity_thr_value, numb_zeros_frst_value, numb_zeros_scnd_value, numb_zeros_thrd_value, slot_strt_frst_value, slot_end_frst_value, slot_strt_scnd_value, slot_end_scnd_value, slot_strt_thrd_value, slot_end_thrd_value):

        parallel_folder  =  analysis_folder + "_SpotsFiltered"
        os.mkdir(parallel_folder)
        print(parallel_folder)
        shutil.copyfile(analysis_folder + '/crop_vect.npy', parallel_folder + '/crop_vect.npy')
        shutil.copyfile(analysis_folder + '/im_red_smpl.npy', parallel_folder + '/im_red_smpl.npy')
        shutil.copyfile(analysis_folder + '/nuclei_tracked.npy', parallel_folder + '/nuclei_tracked.npy')
        shutil.copyfile(analysis_folder + '/nucs_spts_ch.bin', parallel_folder + '/nucs_spts_ch.bin')

        book                   =  load_workbook(analysis_folder + '/journal.xlsx')
        software_version       =  book.worksheets[0]["B24"].value
        gaus_log_detect_value  =  book.worksheets[0]["B2"].value
        param_detect_value     =  book.worksheets[0]["B3"].value
        gfilt_water_var        =  book.worksheets[0]["B4"].value
        circ_thr_var           =  book.worksheets[0]["B5"].value
        dist_thr_var           =  book.worksheets[0]["B6"].value
        spots_thr_var          =  book.worksheets[0]["B7"].value
        volume_thr_var         =  book.worksheets[0]["B8"].value
        start_cut_var          =  book.worksheets[0]["B9"].value
        end_cut_var            =  book.worksheets[0]["B10"].value
        max_dist               =  book.worksheets[0]["B11"].value
        time_step_value        =  book.worksheets[0]["B12"].value
        px_brd                 =  book.worksheets[0]["B13"].value
        time_zero              =  book.worksheets[0]["B15"].value
        pix_size               =  book.worksheets[0]["B16"].value
        pix_size_Z             =  book.worksheets[0]["B17"].value
        t_track_end_value      =  book.worksheets[0]["B18"].value

        nuclei_tracked    =  np.load(analysis_folder + '/nuclei_tracked.npy')
        spots_tracked     =  SpotsConnection.SpotsConnection(nuclei_tracked, np.sign(spots_3D.spots_vol), max_dist).spots_tracked
        nuc_active        =  NucleiSpotsConnection.NucleiSpotsConnection(spots_tracked, nuclei_tracked)

        tifffile.imwrite(str(parallel_folder) + "/false_2colors.tiff", nuc_active.nuclei_active3c.astype("uint16"))
        np.save(parallel_folder + '/spots_3D_tzxy.npy', spots_3D.spots_tzxy.astype("uint16"))
        np.save(parallel_folder + '/spots_3D_vol.npy', spots_3D.spots_vol.astype("uint16"))
        np.save(parallel_folder + '/spots_3D_coords.npy', spots_3D.spots_coords.astype("uint16"))
        np.save(parallel_folder + '/spots_3D_ints.npy', spots_3D.spots_ints.astype("uint16"))
        np.save(parallel_folder + '/spots_tracked.npy', spots_tracked.astype("uint16"))

        idx          =  np.unique(spots_tracked)[1:]
        av_in_spots  =  np.zeros((idx.size, spots_3D.spots_coords[-1, 0]))
        i            =  0
        for k in idx:
            av_in_spots[i, :]  =   ((spots_tracked == k) * spots_3D.spots_ints).sum(2).sum(1)
            i                  +=  1

        ctrs  =  np.zeros((idx.size, spots_tracked.shape[0], 2))
        i     =  0
        for k in idx:
            aa  =  (spots_tracked == k).astype(int)
            for t in range(spots_tracked.shape[0]):
                if aa[t, :, :].sum():
                    aa_rgp         =  regionprops(aa[t, :, :])
                    ctrs[i, t, :]  =  aa_rgp[0]['Centroid'][0], aa_rgp[0]['Centroid'][1]
            i  +=  1

        book    =  xlsxwriter.Workbook(parallel_folder + '/journal.xlsx')                                                                  # write results
        sheet1  =  book.add_worksheet("Sheet1")
        sheet2  =  book.add_worksheet("Sheet2")

        sheet1.write(0, 4, "Time")
        sheet1.write(0, 5, "Frame")
        sheet1.write(0, 6, "Active Nuc")

        for t in range(nuc_active.n_active_vector.size):
            sheet1.write(t + 1, 4, (time_zero + t) * time_step_value)
            sheet1.write(t + 1, 5, t)
            sheet1.write(3 + int(ctrs.shape[1]) + t, 4, (time_zero + t) * time_step_value)
            sheet1.write(3 + int(ctrs.shape[1]) + t, 5, t)
            sheet1.write(5 + 2 * int(ctrs.shape[1]) + t, 4, (time_zero + t) * time_step_value)
            sheet1.write(5 + 2 * int(ctrs.shape[1]) + t, 5, t)
            sheet1.write(t + 1, 6, nuc_active.n_active_vector[t])

        sheet1.write(1, 0, "Detection Algorithm")
        sheet1.write(2, 0, "Detection Parameter")
        sheet1.write(3, 0, "Gaussian Filter W-Shed")
        sheet1.write(4, 0, "Circularity Threshold")
        sheet1.write(5, 0, "Distance Threshold for Nuclei Tracking")
        sheet1.write(6, 0, "Spots Segmentation Threshold")
        sheet1.write(7, 0, "Spots Volume Threshold")
        sheet1.write(8, 0, "First Frame Considered")
        sheet1.write(9, 0, "Last Frame Considered")
        sheet1.write(10, 0, "Spot-Nucleus max distance")
        sheet1.write(11, 0, "Time Step Value")
        sheet1.write(12, 0, "Num Border Pix Removed")
        sheet1.write(14, 0, "Time of First Frame from Mitosis")
        sheet1.write(15, 0, "Pixel Size")
        sheet1.write(16, 0, "Pixel Size Z")
        sheet1.write(17, 0, "Frames Pre-Track")

        sheet1.write(1, 1, gaus_log_detect_value)
        sheet1.write(2, 1, param_detect_value)
        sheet1.write(3, 1, gfilt_water_var)
        sheet1.write(4, 1, circ_thr_var)
        sheet1.write(5, 1, dist_thr_var)
        sheet1.write(6, 1, spots_thr_var)
        sheet1.write(7, 1, volume_thr_var)
        sheet1.write(8, 1, start_cut_var)
        sheet1.write(9, 1, end_cut_var)
        sheet1.write(10, 1, max_dist)
        sheet1.write(11, 1, time_step_value)
        sheet1.write(12, 1, px_brd)
        sheet1.write(14, 1, int(time_zero))
        sheet1.write(15, 1, pix_size)
        sheet1.write(16, 1, pix_size_Z)
        sheet1.write(17, 1, t_track_end_value)

        sheet1.write(18, 0, "File Names")

        for i in range(len(fnames)):
            jj  =  str(fnames[i]).rfind("/")
            sheet1.write(19 + i, 0, str(fnames[i])[jj + 1:])

        sheet1.write(20 + len(fnames), 0, "Date")
        sheet1.write(20 + len(fnames), 1, datetime.datetime.now().strftime("%d-%b-%Y"))
        sheet1.write(21 + len(fnames), 0, "Software Version")
        sheet1.write(21 + len(fnames), 1, software_version)

        for i in range(np.min([av_in_spots.shape[0], 240])):
            sheet1.write(0, i + 7, "Spot_" + str(int(idx[i])))
            for t in range(av_in_spots.shape[1]):
                sheet1.write(t + 1, i + 7, av_in_spots[i, t])

        if av_in_spots.shape[0] > 240:
            for i in range(av_in_spots.shape[0] - 240):
                sheet2.write(0, i + 7, "Spot_" + str(int(idx[i + 240])))
                for t in range(av_in_spots.shape[1]):
                    sheet2.write(t + 1, i + 7, av_in_spots[i + 240, t])

        for i in range(np.min([ctrs.shape[0], 240])):
            sheet1.write(2 + int(ctrs.shape[1]), i + 7, "X coord")
            sheet1.write(4 + 2 * int(ctrs.shape[1]), i + 7, "Y coord")

            for t in range(av_in_spots.shape[1]):
                sheet1.write(t + 3 + int(ctrs.shape[1]), i + 7, ctrs[i, t, 0])
                sheet1.write(t + 5 + 2 * int(ctrs.shape[1]), i + 7, ctrs[i, t, 1])

        if ctrs.shape[0] > 240:
            for i in range(ctrs.shape[0] - 240):
                sheet2.write(2 + int(ctrs.shape[1]), i + 7, "X coord")
                sheet2.write(4 + 2 * int(ctrs.shape[1]), i + 7, "Y coord")

                for t in range(av_in_spots.shape[1]):
                    sheet2.write(t + 3 + int(ctrs.shape[1]), i + 7, ctrs[i + 240, t, 0])
                    sheet2.write(t + 5 + 2 * int(ctrs.shape[1]), i + 7, ctrs[i + 240, t, 1])
        book.close()

        x_vv  =  np.arange(nuc_active.n_active_vector.size)
        plt   =  pg.plot()
        plt.plot(x_vv, nuc_active.n_active_vector, pen='r', title='Number of Active Nuclei')

        WriteSptsIntsDividedByBkg.WriteSptsIntsDividedByBkg(parallel_folder, green4D, spots_3D, spots_tracked)

        filetxt  =  open(parallel_folder + '/filterInfo.txt', "w")
        filetxt.write("solidity_thr_value:  " + str(solidity_thr_value)  + "\n")
        filetxt.write("numb_zeros_frst_value:  " + str(numb_zeros_frst_value)  + "\n")
        filetxt.write("numb_zeros_scnd_value:  " + str(numb_zeros_scnd_value)  + "\n")
        filetxt.write("numb_zeros_thrd_value:  " + str(numb_zeros_thrd_value)  + "\n")
        filetxt.write("slot_strt_frst_value:  " + str(slot_strt_frst_value)  + "\n")
        filetxt.write("slot_end_frst_value:  " + str(slot_end_frst_value)  + "\n")
        filetxt.write("slot_strt_scnd_value:  " + str(slot_strt_scnd_value)  + "\n")
        filetxt.write("slot_end_scnd_value:  " + str(slot_end_scnd_value)  + "\n")
        filetxt.write("slot_strt_thrd_value:  " + str(slot_strt_thrd_value)  + "\n")
        filetxt.write("slot_end_thrd_value:  " + str(slot_end_thrd_value))
        filetxt.close()
